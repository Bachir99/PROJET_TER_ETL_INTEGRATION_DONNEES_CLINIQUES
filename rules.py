from operator import truediv
import sys
import os
import pandas as pd
import re
import datetime
from datetime import datetime, timedelta
import numpy as np
import json
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter,range_boundaries
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension


def get_rules_from_nifi_properties(file_path):
    with open(file_path, 'r') as file:
        for line in file:
            if line.startswith('columns_and_rules='):
                json_str = line.strip().split('=')[1]
                return json.loads(json_str)
    return None

def V_today1(df, column_name, reject_list,rejections_count):
    today = datetime.now()
    rule_name = 'V-Today-1'
    def check_date(row):
        dob = pd.to_datetime(row[column_name], errors='coerce')
        if dob > today:
            row_copy = row.copy()
            row_copy['Rejet'] = rule_name
            reject_list.append(row_copy)
            rejections_count[rule_name] += 1
            return False
        return True
    df = df[df.apply(check_date, axis=1)]
    return df


def V_date_of_birth1(df, column_name,reject_list,rejections_count):
    today = datetime.now()
    max_dob = today - timedelta(days=125 * 365)
    rule_name = 'V-DateOfBirth-1'
    def check_dates(row):
        dob = pd.to_datetime(row[column_name], errors='coerce')
        
        if dob is not pd.NaT and dob < max_dob:
            row_copy = row.copy()
            row_copy['Rejet'] = rule_name
            reject_list.append(row_copy)
            rejections_count[rule_name] += 1
            return False

        return True

    df = df[df.apply(check_dates, axis=1)]

    return df

def V_dateOfDeath(df, column_name,reject_list,rejections_count):
    rule_name = 'V-DateofDeath'
    def check_dates(row):
        dob = pd.to_datetime(row['DateOfBirth'], errors='coerce')
        dod = pd.to_datetime(row[column_name], errors='coerce')
        
        if dod is not pd.NaT and dob is not pd.NaT and dod < dob:
            row_copy = row.copy()
            row_copy['Rejet'] = rule_name
            reject_list.append(row_copy)
            rejections_count[rule_name] += 1
            return False

        return True

    df = df[df.apply(check_dates, axis=1)]

    return df

def D_patientDeceased(row,column_name):
    if pd.isna(row['PatientDeceased']):
        if pd.notna(row[column_name]):
            return 'Oui'
    return row['PatientDeceased']


def T_RemoveLeadingZero_1(row, column_name):
    value = row[column_name]
    if isinstance(value, str) and value.startswith('0'):
        row[column_name] = value.lstrip('0')
    return row[column_name]

def V_NotNull1(row,column_name, reject_list,rejections_count):
    rule_name = 'V-NotNull-1'
    if pd.isna(row[column_name]) or row[column_name] == '':
        row_copy = row.copy()
        row_copy['Rejet'] = rule_name
        reject_list.append(row_copy)
        rejections_count[rule_name] += 1
        return False
    return True

def V_NotNull2(row,column_name, warnings_list, warnings_count):
    rule_name = 'V-NotNull-2'
    if pd.isna(row[column_name]) or row[column_name] == '':
        row_copy = row.copy()
        row_copy['Avertissement'] = rule_name
        warnings_list.append(row_copy)
        warnings_count[rule_name] += 1

def V_length50(row, column_name, warnings_list, warnings_count):
    rule_name = 'V-length50'
    if len(str(row[column_name])) > 50:
        row_copy = row.copy()
        row_copy['Avertissement'] = rule_name
        warnings_list.append(row_copy)
        warnings_count[rule_name] += 1

def V_length100(row, column_name, warnings_list, warnings_count):
    rule_name = 'V-length100'
    if len(str(row[column_name])) > 100:
        row_copy = row.copy()
        row_copy['Avertissement'] = rule_name
        warnings_list.append(row_copy)
        warnings_count[rule_name] += 1

def V_alpha1(row, column_name,reject_list, rejections_count):
    rule_name = 'V-alpha-1'
    if not str(row[column_name]).isalpha():
        row_copy = row.copy()
        row_copy['Rejet'] = rule_name
        reject_list.append(row_copy)
        rejections_count[rule_name] += 1
        return False
    return True

def V_alpha2(row, column_name,warnings_list, warnings_count):
    rule_name = 'V-alpha-2'
    if not str(row[column_name]).isalpha():
        row_copy = row.copy()
        row_copy['Avertissement'] = rule_name
        warnings_list.append(row_copy)
        warnings_count[rule_name] += 1

def D_Null_1(row, column_name):
    value = row[column_name]
    if value == '':
        row[column_name] = 'Null'
    return row[column_name]

def enlever_all_null_colonnes(df):
    for col in df.columns:
        if df[col].isna().all():
            df.drop(col, axis=1, inplace=True)


def deduplicate(input_csv,summary_df):
   
    # Replace 'PatientNumber' with the name of the column you want to use for deduplication
    deduplicated_df = input_csv.drop_duplicates(subset='PatientNumber', keep='first')


    # Find the duplicate rows
    duplicates_df = input_csv[input_csv.duplicated(subset='PatientNumber', keep='first')]

    # Add a column to the duplicates DataFrame indicating that these rows are duplicates
    duplicates_df = duplicates_df.assign(Rejet='Duplication')


    # Compter le nombre de lignes dupliquées
    nb_lignes_dupliquees = len(input_csv) - len(deduplicated_df)

    summary_df = summary_df.append({'Rule': 'Duplication', 'Rejected': nb_lignes_dupliquees, 'Warned': 0}, ignore_index=True)

    return deduplicated_df,summary_df,duplicates_df

def create_excel(lines_df,initial_row_count,warnings_count,rejections_count):
    # Calculate total warnings
    total_warnings = sum(warnings_count.values())

    # Calculate total rejections
    total_rejections = sum(rejections_count.values())
    # Créer un nouveau classeur Excel
    wb = openpyxl.Workbook()

    worksheet = wb.active
    worksheet.title = 'Summary'
    # Merge cells D2:E3 and set the fill color to blue
    cell_range = 'D2:E3'
    worksheet.merge_cells(cell_range)

    fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    for row in worksheet[cell_range]:
        for cell in row:
            cell.fill = fill

    # Write the text "Validation report" in black font
    cell = worksheet.cell(row=2, column=4)
    cell.value = "Validation report"
    cell.font = Font(color='000000', bold=True)

    # Center the text horizontally and vertically
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # Créer une feuille de calcul pour les avertissements
    ws_warnings = wb.create_sheet('Details')

    # Écrire le DataFrame des avertissements dans la feuille de calcul
    for r in dataframe_to_rows(lines_df, index=False, header=True):
        ws_warnings.append(r)


    # Définir le tableau
    table = [
        ['Cluster:', 'CIUM'],
        ['Hopital:', 'Hopital1'],
        ['Servicing department', 'Patient']
    ]

    # Écrire les valeurs dans les cellules appropriées
    for row in range(4, 7):
        for col in range(2, 4):
            cell = worksheet.cell(row=row, column=col)
            cell.value = table[row-4][col-2]

    # Fusionner les cellules B4:C4, B5:C5, B6:C6
    for row in range(4, 7):
        start_col, start_row, end_col, end_row = range_boundaries(f'B{row}:C{row}')
        worksheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)

    # Mettre le texte en gras pour les titres
    title_font = Font(bold=True)

    for row in range(4, 7):
        cell = worksheet.cell(row=row, column=2)
        cell.font = title_font

    # Définir le remplissage pour le fond en vert kaki
    fill = PatternFill(start_color='C2D69B', end_color='C2D69B', fill_type='solid')

    # Appliquer le remplissage et les bordures noires aux cellules B4:C6
    for row in range(4, 7):
        for col in range(2, 4):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = fill
            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'),
                                top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'))

    # Centrer le texte dans les cellules
    alignment = Alignment(horizontal='center', vertical='center')

    for row in range(4, 7):
        for col in range(2, 4):
            cell = worksheet.cell(row=row, column=col)
            cell.alignment = alignment

    # Write "Production date" in bold font in cell D7
    cell = worksheet['D7']
    cell.value = 'Production date'
    cell.font = Font(bold=True)

    # Write today's date in cells F7 to H7
    cell = worksheet['F7']
    cell.value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')


    # Appliquer  les bordures noires aux cellules B4:C6
    for col in range(4, 8):
        cell = worksheet.cell(row=7, column=col)
        if col == 5 or col == 7:
            cell.border = Border(
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'))
        
        elif col == 4 :
            cell.border = Border(left=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'))
        else :
            cell.border = Border(
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'))




    # Écrire les valeurs dans les cellules appropriées
    table = [
        'Total number of initial records:',
        'Number of rejected records:',
        'Number of averted records:'
    ]

    for row in range(14, 17):
            cell = worksheet.cell(row=row, column=2)
            cell.value = table[row-14]

    # Appliquer  les bordures noires aux cellules B4:C6
    for row in range(14, 17):
        for col in range(2,6):
            cell = worksheet.cell(row=row, column=col)
            if col == 3 or col == 4:
                cell.border = Border(
                                top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'))
            
            elif col == 2 :
                cell.border = Border(left=Side(border_style='thin', color='000000'),
                                top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'))
            else :
                cell.border = Border(
                                right=Side(border_style='thin', color='000000'),
                                top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'))


    # Définir les cellules à fusionner
    merge_ranges = ['B14:E14', 'B15:E15', 'B16:E16']

    # Fusionner les cellules
    for cell_range in merge_ranges:
        worksheet.merge_cells(cell_range)

    # Écrire les valeurs dans les cellules appropriées
    table = [
        'Records',
        '%']

    for col in range(6,8):
        cell = worksheet.cell(row=13, column=col)
        cell.value = table[col-6]
        font = Font(color='0070C0', bold=True)
        cell.font = font
        fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
        cell.fill = fill
        cell.alignment = alignment
        # Appliquer  les bordures noires aux cellules B4:C6
    for row in range(13, 17):
        for col in range(6,8):
            cell = worksheet.cell(row=row, column=col)
            cell.border = Border(
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'),
                                top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'))
            
    cell = worksheet.cell(row=14, column=6)
    cell.value = initial_row_count
    cell.alignment = alignment
    cell = worksheet.cell(row=15, column=6)
    cell.value = total_rejections
    cell.alignment = alignment
    cell = worksheet.cell(row=16, column=6)
    cell.value = total_warnings
    cell.alignment = alignment    
    cell = worksheet.cell(row=14, column=7)
    cell.value = '100%'
    cell.alignment = alignment    
    warning_percentage = (total_warnings / initial_row_count) * 100
    rejection_percentage = (total_rejections / initial_row_count) * 100
    cell = worksheet.cell(row=15, column=7)
    cell.value = str(rejection_percentage)+'%'
    cell.alignment = alignment
    cell = worksheet.cell(row=16, column=7)
    cell.value = str(warning_percentage)+'%'
    cell.alignment = alignment

    table = [
        "PPM load field name",
        "Validation Type",
        "Rule ID",
        "Validation rule",
        "Number",
        "%"
    ]

    for col, value in enumerate(table, start=2):
        cell = worksheet.cell(row=19, column=col)
        cell.value = value
        font = Font(color='0070C0', bold=True)
        cell.font = font
        fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
        cell.fill = fill
        cell.alignment = alignment

    
    row = 20

    # Parcourir le dictionnaire warnings_count
    for key, value in warnings_count.items():
        if value != 0:
            worksheet.cell(row=row, column=3).value = 'Rejection'
            worksheet.cell(row=row, column=4).value = key
            worksheet.cell(row=row, column=6).value = value
            row += 1
            for col in range(2,8):
                cell = worksheet.cell(row=row-1, column=col)            
                cell.border = Border(
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'),
                                top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'))
    # Parcourir le dictionnaire rejections_count
    for key, value in rejections_count.items():
        if value != 0:
            worksheet.cell(row=row, column=3).value = 'Warning'
            worksheet.cell(row=row, column=4).value = key
            worksheet.cell(row=row, column=6).value = value
            row += 1
            for col in range(2,8):
                cell = worksheet.cell(row=row-1, column=col)            
                cell.border = Border(
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'),
                                top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'))

    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 20
    worksheet.column_dimensions['F'].width = 20
    worksheet.column_dimensions['G'].width = 20
    worksheet.column_dimensions['H'].width = 20
    worksheet.column_dimensions['I'].width = 20
    worksheet.column_dimensions['J'].width = 20

    wb.save('/home/bachir/Bureau/S8/HAI823I TER/resultats/ValidationReport.xlsx')


def main():

    nifi_properties_path = "/home/bachir/Téléchargements/nifi-1.20.0-bin/nifi-1.20.0/conf/nifi.properties"
    rules={}
    rules = get_rules_from_nifi_properties(nifi_properties_path)
    df = pd.read_csv(sys.stdin)
    initial_row_count = len(df)
    
    summary_df = pd.DataFrame(columns=['Rule', 'Rejected', 'Warned','Initial'])
    summary_df = summary_df.append({'Rule': 'Nbr de lignes initial','Initial' : initial_row_count}, ignore_index=True)

    df, summary_df, duplicates_df = deduplicate(df,summary_df)
    
    warnings_list = []
    reject_list = []

    warnings_count = {"V-length50":0,
                      "V-length100":0,
                      "V-alpha-2":0,
                      "V-NotNull-2":0}
    rejections_count = {"V-NotNull-1":0,
                        "V-alpha-1":0,
                        "V-Today-1":0,
                        "V-DateOfBirth-1":0,
                        "V-DateofDeath":0,
                        }


    validation_functions = {
        "V_today1": V_today1,
        "V_dateOfBirth1": V_date_of_birth1,
        "V_dateOfDeath": V_dateOfDeath,
        "D_patientDeceased": D_patientDeceased,
        "V_NotNull1": V_NotNull1,
        "V_NotNull2": V_NotNull2,
        "V_length50": V_length50,
        "V_length100": V_length100,
        "V_alpha1": V_alpha1,
        "V_alpha2": V_alpha2,
        "T_RemoveLeadingZero_1": T_RemoveLeadingZero_1,
        "D_Null_1": D_Null_1
    }

    for column, functions in rules.items():
        for function_name in functions:
            if function_name in validation_functions:
                function = validation_functions[function_name]
                if function_name == "D_patientDeceased":
                    df['PatientDeceased'] = df.apply(lambda row: function(row, column), axis=1)
                elif function_name in ["T_RemoveLeadingZero_1","D_Null_1"]:
                    df[column] = df.apply(lambda row: function(row, column), axis=1)
                elif function_name in ["V_length50", "V_length100", "V_alpha2","V_NotNull2"]:
                    df.apply(lambda row: function(row, column, warnings_list, warnings_count), axis=1)
                elif function_name in ["V_NotNull1","V_alpha1"] :
                    df = df[df.apply(lambda row: function(row, column,reject_list, rejections_count), axis=1)]
                else:
                    df = function(df, column,reject_list, rejections_count)

    #FileDateCreation
    file_date_creation = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    df['FileDateCreation'].iloc[0:1] = file_date_creation


    for rule, count in rejections_count.items():
        summary_df = summary_df.append({'Rule': rule, 'Rejected': count, 'Warned': 0}, ignore_index=True)

    for rule, count in warnings_count.items():
        if rule in summary_df['Rule'].values:
            summary_df.loc[summary_df['Rule'] == rule, 'Warned'] = count
        else:
            summary_df = summary_df.append({'Rule': rule, 'Rejected': 0, 'Warned': count}, ignore_index=True)

    summary_df.to_csv('/home/bachir/Bureau/S8/HAI823I TER/resultats/report_rules.csv', index=False)

    # Ajouter les lignes avec des avertissements et des rejets au dataframe lines_df pour pouvoir les écrire dans le validation report après
    lines_df = pd.DataFrame()
    warnings_df = pd.DataFrame(warnings_list)
    rejections_df = pd.DataFrame(reject_list)
    lines_df = lines_df.append(duplicates_df,ignore_index=True)
    lines_df = lines_df.append(warnings_df, ignore_index=True)
    lines_df = lines_df.append(rejections_df, ignore_index=True)


    create_excel(lines_df,initial_row_count,warnings_count,rejections_count)


    # Supprimer les colonnes qui contiennent uniquement des valeurs NULL pour toutes les lignes
    enlever_all_null_colonnes(df)




    df.to_csv(sys.stdout, index=False)

if __name__ == "__main__":
    main()
